import csv
import os.path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

class Data: # Every line of the file will be made into data class which we use later

    def __init__(self, code: str, position: str, version: str = None) -> None:
        self.code = code
        self.version = version
        self.position = position

    def getCode(self) -> str:
        return self.code

    def getVersion(self) -> str:
        return self.version

    def getPosition(self) -> str:
        return self.position
    
    
class DataBase: # This where we will hold out data


    def __init__(self, name = None) -> None:
        self.info = []
        self.name = name

    def getName(self) -> str:
        return self.name
    
    def add(self, item: Data) -> None:
        self.info.append(item)

    def getDataBase(self) -> list:
        return self.info
    
    def __add__(self, other):
        if isinstance(other, DataBase):
            self.info.extend(other.getDataBase())
        return self
    
    def getPositions(self) -> list: # This method returns all positsion it has in it its database used for later in magic command
        result = []
        for item in self.info:
            if item.getPosition() not in result:
                result.append(item.getPosition())
        return result


    # When working with CSV files
    def getData(self, NameOfTheFIle: str) -> int:
        # Opens the file.
        with open(NameOfTheFIle, encoding='utf-8-sig',) as csv_file:

            # Saves the file as a csv.reader object and separates the lines in file to lists of strings which were separated by the delimiter.
            csv_reader = list(csv.reader(csv_file, delimiter=";"))
            check = self.DatabaseFiller(csv_reader)
            if check == -1:
                return -1
            else:
                return 1
    # When working for XLSX files
    def getDataWithXlsx(self, NameOfTheFile: str) -> int:

        workbook = load_workbook(NameOfTheFile)
        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        check = self.DatabaseFiller(data)
        if check == -1:
            return -1
        else:
            return 1

    def takeApartWithXlsx(self, NameOfTheFile: str, outputFile: str, byWhat, Delimiter) -> int:

        workbook = load_workbook(NameOfTheFile)
        sheet = workbook.active

        data = []
        takenapart = []
        withoutpos = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        startIndex = self.tableStartingPoint(data)
        if startIndex == -1:
            return 402
        correct_index = data[startIndex].index(byWhat)
        if correct_index == None:
            return 400
        for row in data[startIndex+1:]:
            if all(cell is None or cell == "" for cell in row):# Row is empty
                continue
            if(row[correct_index] == None):
                withoutpos.append(row)
            elif len(row[correct_index].split(Delimiter)) > 1:
                for item in row[correct_index].split(Delimiter):
                    current = row
                    current[correct_index] = item
                    takenapart.append(current)
            else:
                takenapart.append(row)

        takenapart.sort(key=lambda x: x[correct_index])

        result = data[0:startIndex+1] + takenapart + withoutpos
    
        output = Workbook()
        newSheet = output.active

        for row_index, row in enumerate(result, start=1):
            for col_index, value in enumerate(row, start=1):
                cell = result[col_index]
                cell = newSheet.cell(row=row_index, column= col_index)
                cell.value = value


        
        output.save(outputFile)
        return 1


    def DatabaseFiller(self, items: list) -> int:
        startingIndex = self.tableStartingPoint(items)
        NoneCount = 0
        if startingIndex == -1:
            return -1
        if "Ver" in items[startingIndex]:
            versionIndex =  items[startingIndex].index("Ver")
        else:
            versionIndex =  None
        koodIndex = items[startingIndex].index("Code")
        positsioonIndex = items[startingIndex].index("Ref")
        for row in items[startingIndex+1:]:
            if row[koodIndex] == None or row[koodIndex] == "":
                continue
            if row[positsioonIndex] == None:
                item = Data(row[koodIndex], f"None{NoneCount}", row[versionIndex])
                NoneCount += 1
                self.add(item)
            else:
                positsions = row[positsioonIndex].split(",")
                for positsion in positsions:
                    if versionIndex == None:
                        item = Data(row[koodIndex].strip(), positsion.strip())    
                    else:
                        item = Data(row[koodIndex].strip(), positsion.strip(), row[versionIndex])
                    self.add(item)
        return 1


    def tableStartingPoint(self, array: list) -> int:
        for i in range(len(array)):
            if "Code" in array[i]:
                return i
        return -1
    

def takeApartXlsx(file_first: str, output_file: str, delimiter: str, bywhat: str) -> int:
    if not os.path.isfile(file_first):
        return 404
    database = DataBase()
    return database.takeApartWithXlsx(file_first, output_file, bywhat, delimiter)

def magic(file_first: str, file_second: str, output: str) -> int:


    workbook = Workbook()
    sheet = workbook.active


    red = Font(color="FF0000")
    black = Font(color="000000")
    green = PatternFill(start_color="5bba75", end_color="5bba75", fill_type="solid")
    purple = PatternFill(start_color="c95da0", end_color="c95da0", fill_type="solid")
    ending = os.path.basename(file_first).split(".")[-1]
    print(f"ending is {ending}")
    if ending  == "csv":
        file1 = DataBase()
        check = file1.getData(file_first)
        if check == -1:
            return 404
        fileUnchanged1 = file1.getDataBase()


        file2 = DataBase()
        check = file2.getData(file_second)
        if check == -1:
            return 404
        fileUnchanged2 = file2.getDataBase()
    if ending == "xlsx":
        file1 = DataBase()
        check = file1.getDataWithXlsx(file_first)
        if check == -1:
            return 404
        fileUnchanged1 = file1.getDataBase()


        file2 = DataBase()
        check = file2.getDataWithXlsx(file_second)
        if check == -1:
            return 404
        fileUnchanged2 = file2.getDataBase()

    positions = list(set((file1.getPositions() + file2.getPositions())))
    positions.sort()
    rows = [["Code", "Version", "Code", "Version", "Ref"]]


    for position in positions:
        foundFile1 = list(filter(lambda x: x.getPosition() == position, fileUnchanged1))
        foundFile2 = list(filter(lambda x: x.getPosition() == position, fileUnchanged2))
        if len(foundFile1) > 0 and len(foundFile2) > 0:
            row = [foundFile1[0].getCode(), foundFile1[0].getVersion()
                    , foundFile2[0].getCode(), foundFile2[0].getVersion(), foundFile2[0].getPosition()]
        if len(foundFile1) > 0 and len(foundFile2) == 0:
            row = [foundFile1[0].getCode(), foundFile1[0].getVersion(), "", "", foundFile1[0].getPosition()]
        if len(foundFile1) == 0 and len(foundFile2) > 0:
            row = ["", "", foundFile2[0].getCode(), foundFile2[0].getVersion(), foundFile2[0].getPosition()]
        rows.append(row)


    header_row = rows[0]
    for col_num, value in enumerate(header_row, start=1):
        cell = sheet.cell(row=1, column=col_num)
        if col_num == 1 or col_num == 2:
            cell.fill = green
        if col_num == 3 or col_num == 4:
            cell.fill = purple
        cell.value = value

    data = rows[1:]
    for row_num, row in enumerate(data, start=2):
        for col_num, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            if col_num in [1, 3]:  # Columns 1 and 3
                if col_num == 1:
                    cell.fill = green
                    if(row[col_num - 1] != row[col_num + 1]):
                        cell.font = red
                if col_num == 3:
                    cell.fill = purple
                    if(row[col_num - 1] != row[col_num - 3]):
                        cell.font = red
            elif col_num in [2, 4]:  # Columns 2 and 4
                if col_num == 2:
                    cell.fill = green
                    if(row[col_num - 1] != row[col_num + 1]):
                        cell.font = red
                if col_num == 4:
                    cell.fill = purple
                    if(row[col_num - 1] != row[col_num - 3]):
                        cell.font = red
            elif col_num == 5:  # Column 5
                cell.font = black
    
    workbook.save(output)
    return 1

def prepareTable(size: int) -> list:
    if size < 1:
        return []
    result = ".Code.Version" * (size)
    result = result.split(".")
    result.append("Ref")
    return result[1:]

def magicWithMultipleFiles(names: list) -> list:
    endings = []
    databases = []
    positsions = []
    fileIndex = 0
    for name in names:
        ending = os.path.basename(name).split(".")[-1]
        endings.append(ending)

    if len(set(endings)) > 1:
        return -1
    
    for name in names:
        item = DataBase(f"File name")
        item.getDataWithXlsx(name)
        databases.append(item)
        positsions.extend(item.getPositions())
        fileIndex += 1

    positsions = list(set(positsions))
    firstRow = prepareTable(len(names))
    excel = []
    databases.sort(key=lambda x: x.getName())

    for positsion in positsions:
        row = []
        for database in databases:
            tempItem = []
            tempItem = list(filter(lambda x: x.getPosition() == positsion, database.getDataBase()))
            if len(tempItem) > 0:
                item = tempItem[0]
                itemRow = [item.getCode(), item.getVersion()]
            else:
                itemRow = ["",""]
            row.extend(itemRow)
        row.append(positsion)
        excel.append(row)
    excel.sort(key=lambda x: x[-1])
    excel.insert(0, firstRow)
    return excel


def getFileRow(names: list) -> list:
    result = []
    for name in names:
        result.append(name.split('\\')[-1])
        result.append("")
    return result

def checkSameCode(row: list) -> bool:
    row = row[:-1]
    for i in range(2, len(row), 2):
        if row[0] != row[i]:
            return False
    return True


def checkSameVersion(row: list) -> bool:
    for i in range(3, len(row), 2):
        if row[1] != row[i]:
            return False
    return True 


def WriteIntoFileFromMultiple(names: list, output: str) -> int:
    red = Font(color="FF0000")
    black = Font(color="000000")
    green = PatternFill(start_color="5bba75", end_color="5bba75", fill_type="solid")
    purple = PatternFill(start_color="c95da0", end_color="c95da0", fill_type="solid")
    matrix = magicWithMultipleFiles(names)

    workbook = Workbook()
    sheet = workbook.active
    header_rows = getFileRow(names)
    index = 0
    for col_num, value in enumerate(header_rows, start=1):
        cell = sheet.cell(row=2, column=col_num)
        if index == 0 or index == 1:
            cell.fill = green
        if index == 2 or index == 3:
            cell.fill = purple
        index += 1
        if index == 4:
            index = 0
        cell.value = value

    for row_index, row in enumerate(matrix, start=3):
        colorIndex = 0
        versionColorCheck = checkSameCode(row)
        codeColorCheck = checkSameVersion(row)
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = value
            if  colorIndex == 0 or colorIndex == 1:
                cell.fill = green
            if colorIndex == 2 or colorIndex == 3:
                cell.fill = purple
            if (colorIndex == 0 or colorIndex == 2) and not versionColorCheck:
                cell.font = red
            if (colorIndex == 1 or colorIndex == 3) and not codeColorCheck:
                cell.font = red
            colorIndex +=1
            if colorIndex == 4:
                colorIndex = 0
        
    workbook.save(output)
    return 1
    



    

